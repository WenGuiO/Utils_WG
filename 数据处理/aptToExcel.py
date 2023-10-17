import os
from openpyxl import *
import IPy

stat = {"总攻击次数": 0, "高": 0, "中": 0, "低": 0, "客户端IP数": 0, "客户端IP": [], "服务端IP数": 0, "服务端IP": [],
        "攻击类型": []}
all_event = {}  # {告警子类型：（
# {
#    eventlist:{（事件名称，来源IP，目的IP，风险等级，域名，单位名称,攻击次数）}
#    high_level_attack_times = 0
#    middle_level_attack_times = 0
#    low_level_attack_times =0
#    all_attack_times = 0
# }）
# }

# 原始数据列名
col_name = {
    "事件ID": 0, "来源IP": 1, "目的IP": 2, "告警子类型": 3, "事件名称": 4, "目的端口": 5, "安全告警威胁等级": 6,
    "起始时间": 7, "dns请求域名": 8,
    "请求头": 9, "请求响应码": 10, "设备获取报文内容": 11, "原始事件": 12, "单位名称": 13, "设备名称": 14
}
# 新数据列名
new_col_name = {
    "攻击类型": 0, "攻击次数": 1, "来源IP": 2, "目的IP": 3, "单位名称": 4, "平台判定": 5, "人工研判": 6, "备注": 7
}


# 自定义文件路劲

def xlsx_gen():
    filepath = "D:\\onlinnote\\OneDrive\\桌面\\亚运\\数据\\" + input("请输入文件名: ") + ".xlsx"
    if not os.path.isfile(filepath):
        print("文件不存在！")
        exit(0)
    first_Sheet = "1"
    NEW_FILE_NAME = filepath.split(".")[0] + "_new.xlsx"
    wb = load_workbook(filepath)
    sheet = wb[first_Sheet]
    wb_clear = wb.create_sheet(title="过滤后数据")
    wb_clear.append(list(col_name.keys()))

    num = 0
    for row in sheet.rows:
        num = num + 1
        if row[col_name["来源IP"]].value == "来源IP":
            continue
        if "组委会" in row[col_name["设备名称"]].value:
            continue
        if "杭州奥体中心综合训练馆" in row[col_name["单位名称"]].value:
            continue
        if "综合馆" in row[col_name["设备名称"]].value:
            continue
        if row[col_name["来源IP"]].value:
            if row[col_name["来源IP"]].value in IPy.IP("172.16.96.0/23"):
                continue
        if "10.212.81.221" == row[col_name["来源IP"]].value:
            continue
        print(num, row[col_name["事件名称"]].value, row[col_name["来源IP"]].value, row[col_name["目的IP"]].value,
              row[col_name["安全告警威胁等级"]].value, row[col_name["dns请求域名"]].value,
              row[col_name["单位名称"]].value)
        try:
            row[col_name["dns请求域名"]].value = row[col_name["dns请求域名"]].value.split("<")[0]
        except:
            pass
        if row[col_name['告警子类型']].value not in all_event:
            all_event.update({row[col_name["告警子类型"]].value: ''})
            first = {"eventlist": [[row[col_name["事件名称"]].value, 1,
                                    row[col_name["来源IP"]].value, row[col_name["目的IP"]].value,
                                    row[col_name["单位名称"]].value, row[col_name["安全告警威胁等级"]].value,
                                    "低风险", row[col_name["dns请求域名"]].value,
                                    ]],
                     "高": 0,
                     "中": 0,
                     "低": 0,
                     "总攻击次数": 1
                     }
            first[row[col_name["安全告警威胁等级"]].value] = 1
            all_event.update({row[col_name["告警子类型"]].value: first})

        else:
            first = all_event[row[col_name["告警子类型"]].value]
            flag = 1
            for i in first["eventlist"]:  # 事件是否重复
                if i[new_col_name["来源IP"]] == row[col_name["来源IP"]].value:
                    if i[new_col_name["目的IP"]] == row[col_name["目的IP"]].value:
                        if i[new_col_name["攻击类型"]] == row[col_name["事件名称"]].value:
                            if i[new_col_name["平台判定"]] == row[col_name["安全告警威胁等级"]].value:
                                if i[new_col_name["备注"]] == row[col_name["dns请求域名"]].value:
                                    if i[new_col_name["单位名称"]] == row[col_name["单位名称"]].value:
                                        first["eventlist"].remove(i)
                                        i[1] = i[1] + 1
                                        first["eventlist"].append(i)
                                        flag = 0
                                        break
            if flag:
                first["eventlist"].append([row[col_name["事件名称"]].value, 1,
                                           row[col_name["来源IP"]].value, row[col_name["目的IP"]].value,
                                           row[col_name["单位名称"]].value, row[col_name["安全告警威胁等级"]].value,
                                           "低风险", row[col_name["dns请求域名"]].value,
                                           ])

            first[row[col_name["安全告警威胁等级"]].value] = 1 + first[row[col_name["安全告警威胁等级"]].value]
            first["总攻击次数"] = first["总攻击次数"] + 1
            all_event.update({row[col_name["告警子类型"]].value: first})

        stat["总攻击次数"] = stat["总攻击次数"] + 1
        stat[row[col_name["安全告警威胁等级"]].value] = stat[row[col_name["安全告警威胁等级"]].value] + 1
        if row[col_name["来源IP"]].value not in stat["客户端IP"]:
            stat["客户端IP"].append(row[col_name["来源IP"]].value)
            stat["客户端IP数"] = stat["客户端IP数"] + 1
        if row[col_name["目的IP"]].value not in stat["服务端IP"]:
            stat["服务端IP"].append(row[col_name["目的IP"]].value)
            stat["服务端IP数"] = stat["服务端IP数"] + 1

        a_row = []
        for key in col_name.keys():
            a_row.append(str(row[col_name[key]].value))
        wb_clear.append(a_row)

    for i in all_event.keys():
        wb_new = wb.create_sheet(title=i)
        first = dict(all_event[i])
        first["eventlist"].sort(key=lambda x: x[new_col_name["目的IP"]], reverse=False)
        first["eventlist"].sort(key=lambda x: x[new_col_name["攻击次数"]], reverse=True)
        first["eventlist"].sort(key=lambda x: x[new_col_name["攻击类型"]], reverse=False)
        try:
            first["eventlist"].sort(key=lambda x: x[new_col_name["备注"]], reverse=False)
        except:
            pass
        first["eventlist"].sort(key=lambda x: x[new_col_name["来源IP"]], reverse=False)
        first["eventlist"].sort(key=lambda x: x[new_col_name["单位名称"]], reverse=True)
        for j in first["eventlist"]:
            j[5] = j[5] + "风险"
            wb_new.append(list(j))
        stat["攻击类型"].append([i, all_event[i]["总攻击次数"]])

    wb2 = wb.create_sheet(title="攻击汇总")
    max1 = ["", 0]
    max2 = ["", 0]

    for i in all_event.keys():
        wb2.append([i, all_event[i]["总攻击次数"]])
        print(i + "类" + str(all_event[i]["总攻击次数"]) + "条，", end="")
        if all_event[i]["总攻击次数"] > max1[1]:
            max2[0], max2[1] = max1[0], max1[1]
            max1[0], max1[1] = i, all_event[i]["总攻击次数"]
        elif all_event[i]["总攻击次数"] > max2[1]:
            max2[0], max2[1] = i, all_event[i]["总攻击次数"]

    wb2.append(["总攻击次数", stat["总攻击次数"]])
    wb2.append(["", ""])
    wb2.append(["", ""])

    wb2["D1"] = max1[0]
    wb2["E1"] = max1[1]
    wb2["D2"] = max2[0]
    wb2["E2"] = max2[1]

    wb2["G1"] = "客户端IP数"
    wb2["H1"] = stat["客户端IP数"]

    wb2["G2"] = "服务端IP数"
    wb2["H2"] = stat["服务端IP数"]

    wb2["G3"] = "高风险"
    wb2["H3"] = stat["高"]

    wb2["G4"] = "中风险"
    wb2["H4"] = stat["中"]

    wb2["G5"] = "低风险"
    wb2["H5"] = stat["低"]

    wb.save(NEW_FILE_NAME)


if __name__ == '__main__':
    xlsx_gen()
